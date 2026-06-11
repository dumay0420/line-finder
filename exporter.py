"""Excel 輸出"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

COUNTY_COLORS = {
    "苗栗縣": "F0F4C3", "台中市": "FEF9E7", "彰化縣": "E8F5E9",
    "南投縣": "FBE9E7", "雲林縣": "E3F2FD", "嘉義縣": "FCE4EC",
    "嘉義市": "F3E5F5", "台南市": "FFF3E0",
}

def build_excel(results: list[dict], industry: str, job_id: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LINE@ 清單"

    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    cal   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lal   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin  = Side(style="thin", color="CCCCCC")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font = Font(color="0563C1", underline="single")

    headers    = ["編號", "縣市", "商家名稱", "地址", "電話", "LINE@ ID", "LINE 官方帳號頁面"]
    col_widths = [5,      9,     24,       40,    14,    22,        40]

    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = cal; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(results, 2):
        ws.row_dimensions[ri].height = 22
        color = COUNTY_COLORS.get(row.get("county", ""), "FFFFFF")
        rfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        vals = [
            ri - 1,
            row.get("county", ""),
            row.get("name", ""),
            row.get("address", ""),
            row.get("phone", ""),
            row.get("line_id", ""),
            row.get("line_url", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = rfill; c.border = bdr
            c.alignment = cal if ci == 1 else lal
            if ci == 7 and v:
                c.hyperlink = v
                c.font = link_font

    ws.freeze_panes = "A2"

    # 說明頁
    ws2 = wb.create_sheet("說明")
    ws2["A1"] = f"{industry} LINE@ 搜尋清單"
    ws2["A1"].font = Font(bold=True, size=14, color="1B4F72")
    info = [
        ("搜尋日期", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("搜尋行業", industry),
        ("搜尋方法", "Scrapling StealthyFetcher + Google site:page.line.me"),
        ("總計收錄", f"{len(results)} 筆"),
    ]
    for i, (k, v) in enumerate(info, 3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 60

    output_dir = Path("/tmp/line_finder_exports")
    output_dir.mkdir(exist_ok=True)
    path = output_dir / f"{job_id}_{industry}_LINE@清單.xlsx"
    wb.save(str(path))
    return str(path)
